#-*-coding:utf-8-*-
 
import matplotlib.pyplot as plt
from wordcloud import WordCloud, STOPWORDS
import numpy as np
import jieba
import pymysql
from openpyxl import Workbook
from openpyxl import load_workbook
from PIL import Image
# from pyecharts import Bar, Pie

class CommentAnalysis():
    
    def __init__(self, *args, **kwargs):
        
        # 配置
        self.width = 900
        self.height = 1200
        self.max_words = 300
        self.font_path = "msyh.ttf"
        self.max_font_size = 200
        self.random_state = 30
        self.mask = np.array(Image.open("shen.webp"))

        self.stars = []
        self.citys = []
        self.text = ""

        #设置词云
        self.wc = WordCloud(
            background_color = "white", #设置背景颜色
            width = self.width,
            height = self.height,
            mask = self.mask,  #设置背景图片
            max_words = self.max_words, #设置最大显示的字数
            stopwords = self.SetStopWords(), #设置停用词
            font_path = self.font_path,
            max_font_size = self.max_font_size,  #设置字体最大值
            random_state = self.random_state, #设置有多少种随机生成状态，即有多少种配色方案
            )

        # 连接数据库
        self.dbName = "movie"
        self.dbTable = "rich"
        self.dbHost = "localhost"
        self.dbPort = "3306"
        self.dbUser = "root"
        self.dbPwd = "root"

    def LoadCmtFromText(self, filename):
        """
        从txt中读取评论
        """
        text = open(filename,"rb").read()
        #结巴分词
        wordlist = jieba.cut(text, cut_all=False, HMM=False)
        wl = " ".join(wordlist)
        #print(wl)#输出分词之后的txt
        return wl
    
    def ConnectDB(self):
        """
        连接数据库
        """
        try:
            self.db = pymysql.connect(host=self.dbHost,
                                      port=int(self.dbPort),
                                      user=self.dbUser,
                                      password=self.dbPwd,
                                      db=self.dbName,
                                      use_unicode=True,
                                      charset="utf8")
            # print('连接上了')
            self.cursor = self.db.cursor()
            return True
        except Exception:
            # print(u"连接服务器失败!")
            return False


    def LoadCmtFromDB(self):
        """
        从数据库中读取评论
        """
        self.ConnectDB()
        sql = """
            SELECT * FROM rich 
        """
        self.cursor.execute(sql)
        results = self.cursor.fetchall()
        # print(results)
        cmt = ""
        for rows in results:
            cmt += rows[4] + "\n"
        # print(text)
        # 结巴分词
        wordlist = jieba.cut(cmt, cut_all=False, HMM=False)
        wl = " ".join(wordlist)
        return wl

    def LoadCmtFromExcel(self, filename, sheetname):
        """
        从Excel中读取评论
        """
        wb = load_workbook(filename)
        ws = wb[sheetname]
        ws_rows_len = 10000 # 评论总行数
        cmt_column = 5 # 评论所在列
        stars_column = 4 # 评分所在列
        city_column = 3 # 城市所在列
        # 有效行开始
        start_rows = 1
        # 评论
        cmt = ""
        for row in range(start_rows, ws_rows_len+1):
            cmt_data = ws.cell(row=row, column=cmt_column).value
            stars_data = ws.cell(row=row, column=stars_column).value
            city_data = ws.cell(row=row, column=city_column).value
            if cmt_data:
                cmt += cmt_data + "\n"
            if stars_data:
                self.stars.append(stars_data)
            if city_data:
                self.citys.append(city_data)
        # print(cmt)
        # 结巴分词
        wordlist = jieba.cut(cmt, cut_all=False, HMM=False)
        self.text = " ".join(wordlist)

    def SetStopWords(self):
        """
        停词设置
        """
        stopwords = STOPWORDS.copy()
        stopwords.add("电影")
        return stopwords

    def ScoresCount(self, data):
        dict = {}
        for key in data:
            dict[key] = dict.get(key, 0) + 1
        # print(dict)
        return dict

    def Render(self):
        attr = ["0.5分", "1分", "1.5分", "2分", "2.5分", "3分", "3.5分", "4分", "4.5分", "5分"]
        stars_dict = self.ScoresCount(self.stars)
        data = [stars_dict["0.5"], 
                stars_dict["1"],
                stars_dict["1.5"],
                stars_dict["2"],
                stars_dict["2.5"],
                stars_dict["3"],
                stars_dict["3.5"],
                stars_dict["4"],
                stars_dict["4.5"],
                stars_dict["5"]]

        # bar = Bar()
        # bar.add("《西虹市首富》评分", attr, data)
        # bar.render()

        # pie = Pie("饼图-圆环图示例", title_pos='center')
        # pie.add("商品B", attr, data, center=[50, 50], is_random=True,
        # radius=[50, 75], rosetype='area',
        # is_legend_show=False, is_label_show=True)
        # pie.render()


    # 数据处理
    def ProcessData(self):
        # 通过txt读取内容
        # wl = self.LoadCmtFromText("rich.txt")
        # 通过数据库读取内容
        # text = self.LoadCmtFromDB()
        # 通过Excel读取内容
        self.LoadCmtFromExcel("rich.xlsx", "rich")
        self.myword = self.wc.generate(self.text)#生成词云

    # 显示
    def Show(self):
        # 显示Mask
        plt.subplot(1,2,1)
        plt.imshow(self.mask)
        plt.axis("off")
        # 显示词云图
        plt.subplot(1,2,2)
        plt.imshow(self.myword)
        plt.axis("off")
        plt.show()

if __name__ == "__main__":
    cmt = CommentAnalysis()
    cmt.ProcessData()
    cmt.Render()
    cmt.Show()
